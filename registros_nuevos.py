
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager


import requests
import csv
import sys
import os
import time

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.common.exceptions import NoSuchWindowException
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent


from bs4 import BeautifulSoup as bs
import urllib.request as urllib2

import openpyxl

from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import workbook

root= tk.Tk()
print("1")
canvas1 = tk.Canvas(root, width = 300, height = 400, bg = 'lightsteelblue2', relief = 'raised')
canvas1.pack()

print("2")
label1 = tk.Label(root, text='Registros nuevos', bg = 'lightsteelblue2')
label1.config(font=('helvetica', 20))
canvas1.create_window(150, 60, window=label1)

all_names_list = []

print("3")
def getExcel ():
    global read_file
    # print("4")
    # import_file_path = filedialog.askopenfilename()
    # read_file = pd.read_csv (import_file_path, encoding = 'latin-1')

    # path = 'C:/Users/Usuario/Downloads/R. Eventos sometidos  pendientes de aprobación.xlsx'
    path = filedialog.askopenfilename()
    wb = openpyxl.load_workbook(filename = path)

    ws = wb.get_sheet_by_name('Registros nuevos')
    my_list = []
    list_name = []

    for cell in ws['G']:
        # print(cell.value)
        my_list.append(cell.value)

    my_list.remove(my_list[0])


    print(my_list)

    for cell in ws['I']:
        # print(cell.value)
        list_name.append(cell.value)

    list_name.remove(list_name[0])
    global all_names_list
    all_names_list = list_name
    print(all_names_list)


    final_list = []
    i = 0

    # print(len(my_list))

    # print(type(my_list))
    print(my_list)
    print("----------------------------------------------------------------------------------------------------------------------------------------------------------------------")
    while i < len(my_list):
        if my_list[i] == 'N/A':
            print("Para el evento numero", i, "no hay radicado ni llave")
            i = i + 1

        if my_list[i] == None:
            print("Para el evento numero", i, "no hay radicado ni llave")
            i = i + 1

        else:
            print(type(my_list[i]), i, my_list[i])
            new_values1 = my_list[i].split("-")
            final_list.append(new_values1)
            i += 1


    filtered_list = [tup for tup in final_list if len(tup) == 2]


    print(filtered_list)

    # create csv file
    with open (r'lista_radicados_y_llaves_Registros_nuevos.csv', 'w', newline='') as write_file:
        write=csv.writer(write_file)
        write.writerows(filtered_list)

    aList=[]
    print("START ROW")
    with open('lista_radicados_y_llaves_Registros_nuevos.csv', 'r') as f:
        for row in csv.reader(f):
            row = [col.strip() for col in row]
            aList.append(row)
        print(aList)

    # create csv file
    with open (r'lista_radicados_y_llaves_Registros_nuevos.csv', 'w', newline='') as write_file:
        write=csv.writer(write_file)
        write.writerows(aList)

    print(path)
    
print("5")  

def switchBrowseButton_CSV():
    if (browseButton_CSV['state'] == tk.NORMAL):
        browseButton_CSV['state'] = tk.DISABLED
    else:
        browseButton_CSV['state'] = tk.NORMAL

def forLoadButton():
    getExcel()
    switchBrowseButton_CSV()

root.lift()

browseButton_CSV = tk.Button(root, text="      Carga el archivo de Excel     ", command=forLoadButton, bg='green', fg='white', font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 130, window=browseButton_CSV)

def automaticQueries ():
    global read_file
    global all_names_list

    product_name_list = all_names_list
    print(product_name_list)
    
    print("6")
    # export_file_path = filedialog.asksaveasfilename(defaultextension='.json')
    # read_file.to_json (export_file_path)

        #Username and password variables of SANOFI for INVIMA
    USERNAME = 'Genfarregistros'
    PASSWORD = 'Registro2016'

    #Define chromes web driver (located in the folder "login automation", the driver is version 83.0.4103.39)
    # driver = webdriver.Chrome('chromedriver.exe')
    # driver = webdriver.Firefox(executable_path=r'C:/Users/Usuario/Documents/AArchivos salomon/Sanofi/Procesos de automatización/login automation/geckodriver.exe')
    # driver = webdriver.Edge(executable_path=r'C:/Users/Usuario/Documents/AArchivos salomon/Sanofi/Procesos de automatización/login automation/MicrosoftWebDriver.exe')

    def resource_path(relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.dirname(__file__)
        return os.path.join(base_path, relative_path)

    # # driver = webdriver.Edge(resource_path('MicrosoftWebDriver.exe'))
    # driver = webdriver.Chrome(resource_path('chromedriver.exe'))
    messagebox.showinfo("Estado", "Se iniciara el proceso para realizar las consultas") 

    options = Options()
    ua = UserAgent()
    userAgent = ua.random
    print(userAgent)
    options.add_argument(f'user-agent={userAgent}')
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument("--headless")



    # options = chrome_options para proxy
    # chrome_options=options para user agent
    # driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options = options)
    driver = webdriver.Chrome(ChromeDriverManager().install())

    # driver = webdriver.Chrome(resource_path('chromedriver.exe'), chrome_options=options, desired_capabilities=options.to_capabilities())


    #Direct the driver to INVIMA's login URL
    driver.get('https://enlinea.invima.gov.co/rs/login/loginUsuario.jsp')

    #Define and add the user for the input field of the login page
    user_input = driver.find_element_by_name('usuario')
    user_input.send_keys(USERNAME)

    #Define and add the password for the input field of the login page
    password_input = driver.find_element_by_name('clave')
    password_input.send_keys(PASSWORD)

    #Click the login button in the login page
    login_button = driver.find_element_by_name('B1')
    login_button.click()

    #Set the location of the driver to the new popup window of the user, once it's already logged on
    driver.switch_to_window(driver.window_handles[1])

    #Function to go to the query site for the status of the radictions
    driver.implicitly_wait(10)
    goTo_queries = driver.find_element_by_link_text('Consulta estado trámite')
    goTo_queries.click()

    # listado de radicados y llaves

    # DCV worksheet
    # with open('lista_radicados_y_llaves_DCV.csv', newline='') as lista:
    #     reader = csv.reader(lista)
    #     data_DCV = [tuple(row) for row in reader]

    #     #tiene que estar dentro del with
    #     print("TEST")
    #     i = 0
    #     while i < len(data_DCV):
    #         print(data_DCV[i])
    #         i += 1

    # EP worksheet
    with open('lista_radicados_y_llaves_Registros_nuevos.csv', newline='') as lista:
        reader = csv.reader(lista)
        data_EP = [tuple(row) for row in reader]

    #     #tiene que estar dentro del with
    #     print("TEST")
    #     i = 0
    #     while i < len(data_EP):
    #         print(data_EP[i])
    #         i += 1

    # # GX worksheet
    # with open('lista_radicados_y_llaves_GX.csv', newline='') as lista:
    #     reader = csv.reader(lista)
    #     data_GX = [tuple(row) for row in reader]

    #     #tiene que estar dentro del with
    #     print("TEST")
    #     i = 0
    #     while i < len(data_GX):
    #         print(data_GX[i])
    #         i += 1

    # # GZ worksheet
    # with open('lista_radicados_y_llaves_GZ.csv', newline='') as lista:
    #     reader = csv.reader(lista)
    #     data_GZ = [tuple(row) for row in reader]

    #     #tiene que estar dentro del with
    #     print("TEST")
    #     i = 0
    #     while i < len(data_GZ):
    #         print(data_GZ[i])
    #         i += 1

    # # Pasteur worksheet
    # with open('lista_radicados_y_llaves_Pasteur.csv', newline='') as lista:
    #     reader = csv.reader(lista)
    #     data_Pasteur = [tuple(row) for row in reader]

    #     #tiene que estar dentro del with
    #     print("TEST")
    #     i = 0
    #     while i < len(data_Pasteur):
    #         print(data_Pasteur[i])
    #         i += 1

    print("------------------------------------------------------------LISTA-RADICADOS---------------------------------------------------------------------------------------------")

    # print(data_GX[0][0])

    # firstRadication = '20191139845'
    # radicationKey = '533332'

    filename = 'estado_Registros_nuevos.csv'

    csv_writer = csv.writer(open(filename, 'w'))

    csv_writer.writerow(['LOCALIZACION', 'Ingresa', 'Ultimo Mov', 'Termina', 'Tipo Documento', 'Nro Documento', 'Llave', 'Nro radicado'])


    i = 0
    while i < len(data_EP):

        firstRadication = data_EP[i][0]
        radicationKey = data_EP[i][1]
        product_name = product_name_list[i]

        input_first_radication = driver.find_element_by_name('radic')
        input_first_radication.send_keys(firstRadication)

        input_radication_key = driver.find_element_by_name('llave')
        input_radication_key.send_keys(radicationKey)

        search_button = driver.find_element_by_name('enviar')
        search_button.click()
        csv_writer.writerow([' ESTADO EVENTO ', ' CON RADICADO: '+firstRadication+' CON LLAVE: '+radicationKey+' DEL PRODUCTO: '+product_name])

        # newQuery_button = driver.find_element_by_name('Nueva Consulta')
        # newQuery_button.click()

        #Web scraping

        # source = driver.page_source
        # # print(source)
        # # url = requests.get("https://enlinea.invima.gov.co/rs/consultas/constr_encab.jsp")

        # # soup = bs(url.content, 'html.parser')
        # soup = bs(source, 'html.parser')

        # filename = 'estado_GX.csv'

        # csv_writer = csv.writer(open(filename, 'w'))

        # title = soup.find('title')

        # driver.switch_to.frame(‘frame_name’)

        # time.sleep(0.08)
        # driver.implicitly_wait(30)
        # iFrame = driver.find_element_by_xpath('//*[@id="TR2"]/td/iframe')
        # time.sleep(0.05)
        # driver.switch_to.frame(iFrame)
        # time.sleep(0.05)
        # source_iframe = driver.page_source
        # time.sleep(0.5)
        # # print(source_iframe)

        try:
            time.sleep(0.5)
            # driver.implicitly_wait(30)
            print("entered iframe")
            iFrame = driver.find_element_by_xpath('//*[@id="TR2"]/td/iframe')
            driver.switch_to.frame(iFrame)
            print(iFrame)

        except NoSuchWindowException:
            print("couldnt entered iframe")
            continue

        source_iframe = driver.page_source
        time.sleep(0.5)
        # WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "myDynamicElement")))
        soup_iframe = bs(source_iframe, 'html.parser')
        time.sleep(0.05)
        form = soup_iframe.find("form")

        try:

            print("inicio tabla")

            #Tabla datos radicados
            tablaDatos = form.find_next("table")

            #Tabla eventos
            tablaEventos = tablaDatos.find_next("table")
            print(tablaEventos)

            for tr in soup_iframe.find("form").find_next("table").find_next("table").find_all('tr'):
                data = []

                #extract table data

                for td in tr.find_all('td'):
                    data.append(td.text.strip())

                if data:
                    print("Insterting table data: {}".format(','.join(data)))
                    csv_writer.writerow(data)

                # csv_writer.writerow(['ESTADO EVENTO: ', firstRadication])


            driver.switch_to.default_content()
            search_again_button = driver.find_element_by_id('INPUT3')
            search_again_button.click()
            print(firstRadication)
            i += 1

        except:
            # csv_writer.writerow(['Error: No existen datos para este radicado:', firstRadication])

            # driver.switch_to.default_content()
            # time.sleep(0.1)
            # search_again_button = driver.find_element_by_id('INPUT3')
            # search_again_button.click()
            # i += 1
            # traceback.print_exc()
            print("cant enter table")
            csv_writer.writerow(['Error: No existen datos para este radicado:', firstRadication])

            mensaje = 'Hubo un error en la consulta del radico numero ' + firstRadication + ' con llave ' + radicationKey
            # root.lift()
            root.withdraw()
            # messagebox.showinfo("Estado", mensaje) 
            


            driver.switch_to.default_content()
            time.sleep(0.05)
            search_again_button = driver.find_element_by_id('INPUT3')
            search_again_button.click()
            i += 1



    driver.quit()
    messagebox.showinfo("Estado", "Se terminaron de realizar las consultas") 
    root.deiconify()
    root.lift()

print("7")
saveAsButton_Status = tk.Button(root, text='Realizar consultas', command=automaticQueries, bg='green', fg='white', font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 180, window=saveAsButton_Status)

def convertToJSON ():
    global read_file
    
    print("6")

    import_file_path = filedialog.askopenfilename()
    read_file = pd.read_csv (import_file_path, encoding = 'latin-1')
    # export_file_path = filedialog.asksaveasfilename(defaultextension='.json')
    # read_file.to_json (export_file_path)

    wb = openpyxl.Workbook()
    ws = wb.active

    with open(import_file_path) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    wb.save('estado_Registros_nuevos_Excel.xlsx')

    messagebox.showinfo("Estado de la conversion", "Espere mientras se realiza la conversion") 

    wb = openpyxl.load_workbook('estado_Registros_nuevos_Excel.xlsx')
    ws = wb.active
    # ws['A1'].fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    cellCoordinate = ''
    for row in ws.rows:
        if row[0].value == " ESTADO EVENTO ":
            for cell in row:

                # print(cell.value, end=" ")
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                # print(cell.coordinate)
                cellCoordinate = cell.coordinate
                print("La coordenada es: " + cellCoordinate)
                print(ws[cellCoordinate])
                wb.save('estado_Registros_nuevos_Excel.xlsx')

        elif row[0].value == "Error: No existen datos para este radicado:":
            for cell in row:

                # print(cell.value, end=" ")
                cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                # print(cell.coordinate)
                cellCoordinate = cell.coordinate
                print("La coordenada es: " + cellCoordinate)
                print(ws[cellCoordinate])
                wb.save('estado_Registros_nuevos_Excel.xlsx')
                 

    messagebox.showinfo("Estado de la conversion", "Se termino de convertir el documento a Excel, el archivo se llama estado_EP_Excel.xlsx y se encuentra en la misma carpeta que el programa main_program.exe")
    os.remove('estado_Registros_nuevos.csv')
    os.remove('lista_radicados_y_llaves_Registros_nuevos.csv')


print("7")
saveAsButton_JSON = tk.Button(root, text='Convertir a Excel', command=convertToJSON, bg='green', fg='white', font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 230, window=saveAsButton_JSON)

def exitApplication():
    MsgBox = tk.messagebox.askquestion ('Volver al menu','¿Seguro que deseas volver al menu?',icon = 'warning')
    if MsgBox == 'yes':
        root.destroy()
        # import main_program
     
exitButton = tk.Button (root, text='       Volver     ',command=exitApplication, bg='brown', fg='white', font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 280, window=exitButton)

root.mainloop()