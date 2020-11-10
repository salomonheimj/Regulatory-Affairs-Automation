from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

import requests
import csv
import sys
import os
import time
import glob

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.common.exceptions import NoSuchWindowException
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException


from xlsxwriter.workbook import Workbook
from pyexcel.cookbook import merge_all_to_a_book
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# import asposecells
# from asposecells.api import Workbook, FileFormatType

import win32api

import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd

from fake_useragent import UserAgent

from bs4 import BeautifulSoup as bs
import urllib.request as urllib2
import traceback

from datetime import date

today = date.today()

d1 = today.strftime("%d")
d2 = today.strftime("%d/%m/%Y")
# print("d1 =", d1)
day = int(d1) - 1
# print(day)

new_day = str(day)

day_to_test = '23'


driver = webdriver.Chrome(ChromeDriverManager().install())
# driver = webdriver.Chrome(executable_path=r'chromedriver.exe')

driver.get('https://reactivoenlinea.invima.gov.co/TransparenciaWeb/')

driver.find_element_by_xpath('//*[@id="details-button"]').click()

driver.find_element_by_xpath('//*[@id="proceed-link"]').click()

selection_menu = driver.find_element_by_xpath('//*[@id="form1:seleccionGrupo"]/div[3]').click()

driver.find_element_by_xpath('//*[@id="form1:seleccionGrupo_1"]').click()

driver.find_element_by_xpath('//*[@id="form1:popupButtonCal"]/button').click()

from_day = driver.find_element_by_xpath("//td[not(contains(@class,'ui-datepicker-div'))]/a[text()='" + day_to_test + "']").click()

driver.find_element_by_xpath('//*[@id="form1:j_idt22"]').click()

# Bucle for retrying to scrap from the errors of the loading times at the Invima web app
for i in range(0,100):
	while True:

		# Try to scrap and create the corresponding Excel file with the scraped table data
		try:
			time.sleep(0.5)
			# Check if the current day of the query has information to scrap
			# If it doesn't, it will display a tkinter message box
			try:
				time.sleep(0.5)
				driver.find_element_by_xpath('//*[@id="form1:messages"]/div/span')
				win32api.MessageBox(0, 'No se han encontrado resultados para este dia', 'Estado de la consulta en el Invima', 0x00001000) 
				print("NO HAY INFO")

			# If it does, it will scrap its content and create the CSV and then Excel file
			except NoSuchElementException:
			    print("SI HAY INFO")


			    filename = 'sometimientos.csv'

			    f = open(filename, 'w')

			    csv_writer = csv.writer(f)

			    csv_writer.writerow(['Titular', 'Direccion', 'Expediente', 'Num Radicado', 'Fecha Solicitud', 'Tipo Tramite', 'Principio Activo', 'Cantidad', 'Unidad', 'Modalidad', 'Fabricante', 'Dir Fabricante', 'Importador', 'Dir Importador', 'Radicación'])

			    source = driver.page_source
			    soup_frame = bs(source, 'html.parser')
			    form = soup_frame.find("form")
			    # print(form)
			    time.sleep(2)

			    tabla1 = soup_frame.findAll("table", {"role": "grid"})[0].find_next("tbody")

			    time.sleep(1)
			    # print(tabla1)

			    for tr in soup_frame.findAll("table", {"role": "grid"})[0].find_next("tbody").findAll("tr"):
			    	data = []
			    	for td in tr.findAll("td"):

					    # print(td.text.strip)
					    data.append(td.text.strip())
					    print("finding info")

			    	if data:
			    		print("Insterting table data: {}".format(', '.join(data)))
			    		print("wrote data")
			    		csv_writer.writerow(data)

			# The CSV file opened to write the bs scraping results have to closed, so it can save and be transformed into xlsx format
			f.close()

			read_file = pd.read_csv('sometimientos.csv', encoding = 'latin-1')
			wb = openpyxl.Workbook()
			ws = wb.active
			with open('sometimientos.csv') as f:
				reader = csv.reader(f, delimiter=',')
				for row in reader:
					ws.append(row)

			for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
				for cell in rows:
					cell.fill = PatternFill(start_color='DBF4FF', end_color='DBF4FF', fill_type="solid")


			wb.save('sometimiento de medicamentos - Excel.xlsx')
			os.remove('sometimientos.csv')
			driver.quit()
		except IndexError:
			continue
		break
